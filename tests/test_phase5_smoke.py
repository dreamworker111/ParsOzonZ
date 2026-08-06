import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock, patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    from PyQt6.QtCore import Qt
    from PyQt6.QtWidgets import QApplication
except ImportError as exc:  # pragma: no cover - environment without GUI deps
    Qt = None
    QApplication = None
    _GUI_IMPORT_ERROR = exc
else:
    _GUI_IMPORT_ERROR = None

from ozon_parser.config import MOBILE_MODE
from ozon_parser.export import HEADERS, ProductRow, export_products
from ozon_parser.filters import FilterOptionNode
from ozon_parser.login import MobileBrowserLoginSession

if _GUI_IMPORT_ERROR is None:
    from ozon_parser.app import FilterTreeWidget, MainWindow, scaled_font_size


@unittest.skipIf(_GUI_IMPORT_ERROR is not None, f"GUI deps unavailable: {_GUI_IMPORT_ERROR}")
class GuiModeSmokeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def test_mobile_auth_controls_and_missing_session_guard(self):
        window = MainWindow()
        window.show()
        self.app.processEvents()
        try:
            self.assertTrue(window.chrome_mode_label.isVisible())
            self.assertFalse(window.mobile_login_btn.isVisible())

            window.browser_mode_combo.setCurrentIndex(
                window.browser_mode_combo.findData(MOBILE_MODE)
            )
            window.auth_mode_combo.setCurrentIndex(
                window.auth_mode_combo.findData(True)
            )
            self.app.processEvents()

            self.assertFalse(window.chrome_mode_label.isVisible())
            self.assertTrue(window.mobile_login_btn.isVisible())
            with (
                patch(
                    "ozon_parser.app.has_mobile_saved_session",
                    return_value=False,
                ),
                patch("ozon_parser.app.QMessageBox.warning") as warning,
            ):
                self.assertFalse(window._validate_mobile_auth())
            warning.assert_called_once()
            self.assertIn("завершите вход", warning.call_args.args[2].lower())
        finally:
            window.close()

    def test_compact_screen_font_scaling_has_no_threshold_jump(self):
        self.assertLessEqual(
            scaled_font_size(800) - scaled_font_size(799),
            1,
        )
        self.assertEqual(scaled_font_size(1080), 27)

    def test_help_button_and_instructions_content(self):
        from ozon_parser.instructions import build_user_instructions

        window = MainWindow()
        self.app.processEvents()
        try:
            self.assertEqual(window.help_btn.text(), "Инструкция")
            self.assertTrue(callable(window._show_instructions))
            text = build_user_instructions().lower()
            self.assertIn("chrome", text)
            self.assertIn("загрузить категории", text)
            self.assertIn("загрузить подкатегории", text)
            self.assertIn("parse_checkpoint", text)
            self.assertIn("большой объём", text)
        finally:
            window.close()

    def test_category_can_be_renamed_without_changing_ozon_target(self):
        tree = FilterTreeWidget()
        child = FilterOptionNode(
            id="child-id",
            name="Старое имя",
            url="https://www.ozon.ru/category/123/",
            param_key="category",
            param_value="123",
            category_id="123",
            category_name="Старое имя",
            parent_name="Родитель",
        )
        tree.populate_categories([child])
        item = tree.topLevelItem(0)

        self.assertTrue(tree.rename_category_item(item, "  Новое   имя  "))
        item.setCheckState(0, Qt.CheckState.Checked)
        target = tree.selected_leaf_categories()[0]

        self.assertEqual(item.text(0), "Новое имя")
        self.assertEqual(target.name, "Родитель → Новое имя")
        self.assertEqual(target.category_name, "Новое имя")
        self.assertEqual(target.id, "child-id")
        self.assertEqual(target.category_id, "123")
        self.assertEqual(target.url, "https://www.ozon.ru/category/123/")

    def test_category_tree_shows_hierarchy_affiliation(self):
        tree = FilterTreeWidget()
        leaf = FilterOptionNode(
            id="leaf",
            name="Android",
            url="https://www.ozon.ru/category/3/",
            param_key="category",
            param_value="3",
            category_id="3",
            category_name="Android",
            parent_name="Смартфоны",
        )
        mid = FilterOptionNode(
            id="mid",
            name="Смартфоны",
            url="https://www.ozon.ru/category/2/",
            param_key="category",
            param_value="2",
            category_id="2",
            category_name="Смартфоны",
            parent_name="Электроника",
            children=[leaf],
        )
        root = FilterOptionNode(
            id="root",
            name="Электроника",
            url="https://www.ozon.ru/category/1/",
            param_key="category",
            param_value="1",
            category_id="1",
            category_name="Электроника",
            children=[mid],
        )
        tree.populate_categories([root])

        root_item = tree.topLevelItem(0)
        mid_item = root_item.child(0)
        leaf_item = mid_item.child(0)

        self.assertEqual(root_item.text(0), "Электроника  (1)")
        self.assertTrue(root_item.text(0).startswith("Электроника"))
        self.assertIn("Смартфоны", mid_item.text(0))
        self.assertIn("(1)", mid_item.text(0))
        self.assertIn("Android", leaf_item.text(0))
        self.assertIn("Смартфоны", leaf_item.text(0))
        self.assertEqual(
            leaf_item.data(0, tree.ROLE_FULL_PATH),
            "Электроника → Смартфоны → Android",
        )

        leaf_item.setCheckState(0, Qt.CheckState.Checked)
        target = tree.selected_leaf_categories()[0]
        self.assertEqual(target.name, "Электроника → Смартфоны → Android")
        self.assertEqual(target.parent_name, "Смартфоны")

    def test_checked_root_with_unchecked_children_does_not_parse_root(self):
        """Regression: after subcategory load, root stays checked but kids unchecked.

        Parsing must not fall back to the whole root (e.g. Electronics).
        """
        tree = FilterTreeWidget()
        leaf = FilterOptionNode(
            id="leaf",
            name="Заколки",
            url="https://www.ozon.ru/category/3/",
            param_key="category",
            param_value="3",
            category_id="3",
            category_name="Заколки",
        )
        root = FilterOptionNode(
            id="root",
            name="Электроника",
            url="https://www.ozon.ru/category/1/",
            param_key="category",
            param_value="1",
            category_id="1",
            category_name="Электроника",
            children=[leaf],
        )
        tree.populate_categories([root])
        root_item = tree.topLevelItem(0)
        leaf_item = root_item.child(0)

        # Simulate inconsistent state after blocked-signal subcategory restore.
        tree._block_signals = True
        root_item.setCheckState(0, Qt.CheckState.Checked)
        leaf_item.setCheckState(0, Qt.CheckState.Unchecked)
        tree._block_signals = False

        self.assertEqual(tree.selected_leaf_categories(), [])

        leaf_item.setCheckState(0, Qt.CheckState.Checked)
        selected = tree.selected_leaf_categories()
        self.assertEqual(len(selected), 1)
        self.assertEqual(selected[0].category_id, "3")
        self.assertIn("Заколки", selected[0].name)

    def test_finalize_catalog_load_clears_checks_for_explicit_leaf_pick(self):
        tree = FilterTreeWidget()
        tree.set_initial_roots(
            [
                FilterOptionNode(
                    id="root",
                    name="Электроника",
                    url="https://www.ozon.ru/category/1/",
                    param_key="category",
                    param_value="1",
                    category_id="1",
                    category_name="Электроника",
                )
            ],
            pending_subcategories=True,
        )
        root_item = tree.topLevelItem(0)
        root_item.setCheckState(0, Qt.CheckState.Checked)

        tree.finalize_catalog_load(
            [
                FilterOptionNode(
                    id="root",
                    name="Электроника",
                    url="https://www.ozon.ru/category/1/",
                    param_key="category",
                    param_value="1",
                    category_id="1",
                    category_name="Электроника",
                    children=[
                        FilterOptionNode(
                            id="leaf",
                            name="Смартфоны",
                            url="https://www.ozon.ru/category/2/",
                            param_key="category",
                            param_value="2",
                            category_id="2",
                            category_name="Смартфоны",
                        )
                    ],
                )
            ]
        )
        root_item = tree.topLevelItem(0)
        self.assertEqual(root_item.checkState(0), Qt.CheckState.Unchecked)
        self.assertEqual(tree.selected_leaf_categories(), [])
        leaf = root_item.child(0)
        leaf.setCheckState(0, Qt.CheckState.Checked)
        self.assertEqual(tree.selected_leaf_categories()[0].category_id, "2")

    def test_selected_leaves_ignore_ancestor_ids_relisted_as_children(self):
        """Junk parent nodes under a deep branch must not be parsed."""
        tree = FilterTreeWidget()
        leaf_pins = FilterOptionNode(
            id="pins",
            name="Заколки",
            url="https://www.ozon.ru/category/17033/",
            param_key="category",
            param_value="17033",
            category_id="17033",
            category_name="Заколки",
        )
        leaf_combs = FilterOptionNode(
            id="combs",
            name="Гребни",
            url="https://www.ozon.ru/category/17034/",
            param_key="category",
            param_value="17034",
            category_id="17034",
            category_name="Гребни",
        )
        junk_parent = FilterOptionNode(
            id="junk",
            name="Женские аксессуары",
            url="https://www.ozon.ru/category/17000/",
            param_key="category",
            param_value="17000",
            category_id="17000",
            category_name="Женские аксессуары",
        )
        hair = FilterOptionNode(
            id="hair",
            name="Аксессуары для волос",
            url="https://www.ozon.ru/category/17047/",
            param_key="category",
            param_value="17047",
            category_id="17047",
            category_name="Аксессуары для волос",
            children=[leaf_pins, leaf_combs, junk_parent],
        )
        women = FilterOptionNode(
            id="women",
            name="Женские аксессуары",
            url="https://www.ozon.ru/category/17000/",
            param_key="category",
            param_value="17000",
            category_id="17000",
            category_name="Женские аксессуары",
            children=[hair],
        )
        tree.populate_categories([women])

        women_item = tree.topLevelItem(0)
        hair_item = women_item.child(0)
        pins_item = hair_item.child(0)
        combs_item = hair_item.child(1)
        junk_item = hair_item.child(2)

        pins_item.setCheckState(0, Qt.CheckState.Checked)
        combs_item.setCheckState(0, Qt.CheckState.Checked)
        junk_item.setCheckState(0, Qt.CheckState.Checked)

        selected = tree.selected_leaf_categories()
        ids = sorted(t.category_id for t in selected)
        self.assertEqual(ids, ["17033", "17034"])
        self.assertTrue(all("/category/1703" in (t.url or "") for t in selected))

    def test_parse_mode_combo_controls_link_input(self):
        from ozon_parser.config import PARSE_MODE_GLOBAL_CATEGORIES, PARSE_MODE_SELLER_FULL

        window = MainWindow()
        try:
            self.assertEqual(window._selected_parse_mode(), PARSE_MODE_GLOBAL_CATEGORIES)
            self.assertFalse(window.seller_input.isEnabled())
            self.assertEqual(window.seller_input.text(), "")
            self.assertEqual(window.max_products.maximum(), 100000)

            for index in range(window.parse_mode_combo.count()):
                if window.parse_mode_combo.itemData(index) == PARSE_MODE_SELLER_FULL:
                    window.parse_mode_combo.setCurrentIndex(index)
                    break
            self.app.processEvents()
            self.assertTrue(window.seller_input.isEnabled())
            self.assertIn("/seller/", window.seller_input.placeholderText())

            for index in range(window.parse_mode_combo.count()):
                if window.parse_mode_combo.itemData(index) == PARSE_MODE_GLOBAL_CATEGORIES:
                    window.parse_mode_combo.setCurrentIndex(index)
                    break
            self.app.processEvents()
            self.assertFalse(window.seller_input.isEnabled())
            self.assertIn("общий каталог", window.seller_input.placeholderText().lower())
        finally:
            window.close()


class BonusOnlyParsingTests(unittest.TestCase):
    def test_process_card_skips_products_without_bonus(self):
        from ozon_parser.parser import OzonParser, ParseSettings, _ParseState

        parser = OzonParser()
        settings = ParseSettings(
            seller_url="",
            categories=[],
            min_price=None,
            max_price=None,
            max_products=100,
            use_auth=False,
            import_browser_session=False,
        )
        state = _ParseState()
        card = {
            "href": "https://www.ozon.ru/product/item-1/",
            "name": "Товар без баллов",
            "text": "Товар без баллов\n1 000 ₽",
            "html": "<div>Товар без баллов 1000 ₽</div>",
        }

        self.assertIsNone(parser._process_card(card, settings, state))

    def test_process_card_keeps_products_with_bonus(self):
        from ozon_parser.parser import OzonParser, ParseSettings, _ParseState

        parser = OzonParser()
        settings = ParseSettings(
            seller_url="",
            categories=[],
            min_price=None,
            max_price=None,
            max_products=100,
            use_auth=False,
            import_browser_session=False,
        )
        state = _ParseState()
        card = {
            "href": "https://www.ozon.ru/product/item-2/",
            "name": "Товар с баллами",
            "text": "Товар с баллами\n1 200 ₽\n150 баллов за отзыв",
            "html": "<div>150 баллов за отзыв</div>",
        }

        product = parser._process_card(card, settings, state)
        self.assertIsNotNone(product)
        self.assertEqual(product.bonus_points, 150)
        self.assertEqual(product.name, "Товар с баллами")

    def test_process_card_keeps_ruble_bonus_format(self):
        from ozon_parser.parser import OzonParser, ParseSettings, _ParseState

        parser = OzonParser()
        settings = ParseSettings(
            seller_url="",
            categories=[],
            min_price=None,
            max_price=None,
            max_products=100,
            use_auth=False,
            import_browser_session=False,
        )
        state = _ParseState()
        card = {
            "href": "https://www.ozon.ru/product/item-3/",
            "name": "Товар с рублями",
            "text": "Товар с рублями\n2 500 ₽\n200 ₽ за отзыв",
            "html": "<div>200 ₽ за отзыв</div>",
        }

        product = parser._process_card(card, settings, state)
        self.assertIsNotNone(product)
        self.assertEqual(product.bonus_points, 200)
        self.assertEqual(product.name, "Товар с рублями")

    def test_process_card_keeps_all_products_when_bonus_filter_off(self):
        from ozon_parser.parser import OzonParser, ParseSettings, _ParseState

        parser = OzonParser()
        settings = ParseSettings(
            seller_url="",
            categories=[],
            min_price=None,
            max_price=None,
            max_products=100,
            use_auth=False,
            import_browser_session=False,
            bonus_only=False,
        )
        state = _ParseState()
        card = {
            "href": "https://www.ozon.ru/product/item-4/",
            "name": "Обычный товар",
            "text": "Обычный товар\n900 ₽",
            "html": "<div>Обычный товар 900 ₽</div>",
        }

        product = parser._process_card(card, settings, state)
        self.assertIsNotNone(product)
        self.assertEqual(product.bonus_points, 0)
        self.assertEqual(product.name, "Обычный товар")
        self.assertEqual(product.price_discounted, 900.0)


class MobileLoginFlowTests(unittest.TestCase):
    def test_login_verifies_session_after_manual_confirmation(self):
        page = Mock()
        page.url = "https://m.ozon.ru/"
        page.evaluate.return_value = ""

        class FakeContext:
            ready = False

            @property
            def pages(self):
                return [page]

            def cookies(self):
                if not self.ready:
                    return []
                return [
                    {
                        "name": "__Secure-access-token",
                        "value": "present",
                        "domain": ".ozon.ru",
                        "expires": -1,
                    },
                    {
                        "name": "__Secure-user-id",
                        "value": "12345",
                        "domain": ".ozon.ru",
                        "expires": -1,
                    },
                ]

            def close(self):
                pass

        context = FakeContext()
        playwright_manager = Mock()
        playwright_manager.__enter__ = Mock(return_value=object())
        playwright_manager.__exit__ = Mock(return_value=False)

        def confirm() -> bool:
            context.ready = True
            page.url = "https://m.ozon.ru/my/main"
            page.evaluate.return_value = "Выйти Мои заказы"
            return True

        with (
            patch(
                "ozon_parser.login.sync_playwright",
                return_value=playwright_manager,
            ),
            patch(
                "ozon_parser.login.create_persistent_context",
                return_value=(context, page),
            ) as create_context,
            patch("ozon_parser.login.mobile_browser_profile_dir") as profile_dir,
            patch("ozon_parser.login.mark_mobile_session_saved") as mark_saved,
            patch("ozon_parser.login.settle_mobile_login"),
        ):
            self.assertTrue(
                MobileBrowserLoginSession().run(
                    wait_for_confirmation=confirm,
                )
            )

        create_context.assert_called_once_with(
            playwright_manager.__enter__.return_value,
            headless=False,
            user_data_dir=profile_dir.return_value,
            browser_mode=MOBILE_MODE,
        )
        mark_saved.assert_called_once_with()

    def test_login_keeps_asking_until_cancelled_without_session(self):
        page = Mock()
        page.url = "https://m.ozon.ru/"
        page.evaluate.return_value = "Войти или зарегистрироваться"
        context = Mock()
        context.pages = [page]
        context.cookies.return_value = []
        playwright_manager = Mock()
        playwright_manager.__enter__ = Mock(return_value=object())
        playwright_manager.__exit__ = Mock(return_value=False)

        confirms = {"n": 0}

        def wait_for_confirmation() -> bool:
            confirms["n"] += 1
            return confirms["n"] < 3

        with (
            patch(
                "ozon_parser.login.sync_playwright",
                return_value=playwright_manager,
            ),
            patch(
                "ozon_parser.login.create_persistent_context",
                return_value=(context, page),
            ),
            patch("ozon_parser.login.mark_mobile_session_saved") as mark_saved,
            patch("ozon_parser.login.settle_mobile_login"),
            patch("ozon_parser.login._wait_until_authenticated", return_value=False),
        ):
            self.assertFalse(
                MobileBrowserLoginSession().run(
                    wait_for_confirmation=wait_for_confirmation,
                )
            )

        self.assertEqual(confirms["n"], 3)
        mark_saved.assert_not_called()


@unittest.skipIf(load_workbook is None, "openpyxl unavailable")
class XlsxRegressionTests(unittest.TestCase):
    def test_export_contains_all_seven_populated_product_fields(self):
        products = [
            ProductRow(
                name="Товар 2",
                price_discounted=200,
                price_original=250,
                url="https://www.ozon.ru/product/item-2/",
                bonus_points=20,
            ),
            ProductRow(
                name="Товар 1",
                price_discounted=100,
                price_original=125,
                url="https://www.ozon.ru/product/item-1/",
                bonus_points=10,
            ),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = export_products(products, output_dir=Path(temp_dir))
            workbook = load_workbook(path, data_only=False)
            sheet = workbook["Товары"]

            self.assertEqual(
                [sheet.cell(1, column).value for column in range(1, 8)],
                list(HEADERS),
            )
            self.assertEqual(sheet.max_row, 3)
            for row in range(2, 4):
                values = [
                    sheet.cell(row, column).value
                    for column in range(1, 8)
                ]
                self.assertTrue(all(value is not None for value in values))
                self.assertTrue(sheet.cell(row, 6).hyperlink.target)
            self.assertEqual(sheet.cell(2, 1).value, "Товар 1")


if __name__ == "__main__":
    unittest.main()
