from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .config import OUTPUT_DIR
from .utils import price_diff


@dataclass
class ProductRow:
    name: str
    price_discounted: float
    price_original: float
    url: str
    bonus_points: int

    @property
    def diff(self) -> float:
        d, _ = price_diff(self.price_original, self.price_discounted)
        return d

    @property
    def diff_percent(self) -> float:
        _, p = price_diff(self.price_original, self.price_discounted)
        return p


@dataclass
class ExportMeta:
    seller_url: str
    min_price: float | None
    max_price: float | None
    max_products: int
    categories: str = "Все"
    parse_duration: str = "—"
    section_timings: str = "—"


HEADERS = (
    "Название товара",
    "Стоимость со скидкой",
    "Стоимость без скидки",
    "Разница в цене",
    "% разницы",
    "Ссылка на товар",
    "Количество бонусов",
)


def ensure_output_dir() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR


def build_filename() -> str:
    now = datetime.now()
    date_part = now.strftime("%d.%m.%Y")
    time_part = now.strftime("%H_%M")
    return f"Ozon баллы {date_part} {time_part}.xlsx"


def export_products(
    products: list[ProductRow],
    meta: ExportMeta | None = None,
    output_dir: Path | None = None,
) -> Path:
    folder = output_dir or ensure_output_dir()
    folder.mkdir(parents=True, exist_ok=True)

    sorted_products = sorted(
        products,
        key=lambda product: (product.price_discounted, product.name.casefold()),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"

    header_font = Font(name="Roboto", size=11, bold=True)
    cell_font = Font(name="Roboto", size=11)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, product in enumerate(sorted_products, start=2):
        values = [
            product.name,
            product.price_discounted,
            product.price_original,
            product.diff,
            product.diff_percent,
            product.url,
            product.bonus_points,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = Alignment(vertical="center")

        ws.cell(row=row_idx, column=2).number_format = '#,##0.00 "₽"'
        ws.cell(row=row_idx, column=3).number_format = '#,##0.00 "₽"'
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00 "₽"'
        ws.cell(row=row_idx, column=5).number_format = '0.00"%"'
        link_cell = ws.cell(row=row_idx, column=6)
        link_cell.hyperlink = product.url
        link_cell.style = "Hyperlink"

    _autosize_columns(ws, len(HEADERS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(sorted_products) + 1}"

    if meta:
        ws_info = wb.create_sheet("Фильтры", 0)
        info_font = Font(name="Roboto", size=11)
        info_bold = Font(name="Roboto", size=11, bold=True)
        rows = [
            ("Параметр", "Значение"),
            ("Магазин", meta.seller_url),
            ("Фильтры", meta.categories),
            ("Цена от", meta.min_price if meta.min_price is not None else "—"),
            ("Цена до", meta.max_price if meta.max_price is not None else "—"),
            ("Количество товаров", meta.max_products),
            ("Записано товаров", len(sorted_products)),
            ("Время парсинга", meta.parse_duration),
            ("Время по разделам", meta.section_timings),
            ("Дата выгрузки", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ]
        for r, (label, value) in enumerate(rows, start=1):
            c1 = ws_info.cell(row=r, column=1, value=label)
            c2 = ws_info.cell(row=r, column=2, value=value)
            c1.font = info_bold if r == 1 else info_font
            c2.font = info_font
            c1.alignment = Alignment(vertical="center", wrap_text=True)
            c2.alignment = Alignment(vertical="center", wrap_text=True)
        ws_info.column_dimensions["A"].width = 24
        ws_info.column_dimensions["B"].width = 70

    filepath = (folder / build_filename()).resolve()
    wb.save(filepath)
    return filepath


def _autosize_columns(ws, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        # Excel ограничивает ширину значением 255. До этого код обрезал
        # названия и URL на 80 символах, хотя пользователь просит видеть их полностью.
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 255)
